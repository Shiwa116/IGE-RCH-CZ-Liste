"""
Erstellt Reverse Charge and IGE.xlsx aus 3216.xlsx + 3226.xlsx
Kreditoren.xlsx wird bevorzugt aus Y:/HRV/ZZ_GK Tools/ geladen;
falls nicht erreichbar, kann sie manuell ausgewaehlt werden.
"""

import os
import re
import sys
import urllib.request
from datetime import datetime, timedelta
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KREDITOREN_PATH = r"Y:\HRV\ZZ_GK Tools\Kreditoren.xlsx"

# ── Farben ────────────────────────────────────────────────────────────────
_FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")   # dunkelblau
_FILL_ZEBRA_E = PatternFill("solid", fgColor="DCE6F1")   # hellblau (gerade Zeilen)
_FILL_ZEBRA_O = PatternFill("solid", fgColor="FFFFFF")   # weiss (ungerade Zeilen)
_FILL_WARN    = PatternFill("solid", fgColor="FFC7CE")   # rot – Service Date frueherer Monat
_FILL_CALC    = PatternFill("solid", fgColor="FFF2CC")   # hellgold – errechnete CZK-Spalten
_BORDER_CALC_COLOR = "BF8F00"                            # dunkelgold – Rahmen um die CZK-Spalten
_FONT_HEADER  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_FONT_BODY    = Font(name="Calibri", size=10)
_NUM_FORMAT   = '#,##0.00'
_NUM_COLS     = {"Fremdwährung", "Fremdwaehrung", "Betrag", "Betrag CZK"}  # Spalten als Zahl formatieren
_CALC_COLS    = {"Betrag CZK", "Währung CZK"}  # optisch abgesetzte, errechnete Spalten


def _apply_formatting(ws, df):
    """Zebrastreifen, Zahlenformat und Service-Date-Warnung auf ws anwenden."""
    headers = [c.value for c in ws[1]]
    col_idx = {name: i + 1 for i, name in enumerate(headers) if name}

    sd_col    = col_idx.get("Service Date")
    datum_col = col_idx.get("Datum")
    num_cols  = [col_idx[h] for h in headers if h in _NUM_COLS and h in col_idx]
    calc_cols = [col_idx[h] for h in headers if h in _CALC_COLS and h in col_idx]
    calc_min  = min(calc_cols) if calc_cols else None
    calc_max  = max(calc_cols) if calc_cols else None

    thin  = Side(style="thin", color="AAAAAA")
    thick = Side(style="medium", color=_BORDER_CALC_COLOR)

    def cell_border(col, row_idx):
        """Normaler duenner Rahmen, an den Aussenkanten des CZK-Blocks dicker/goldfarben."""
        left   = thick if col == calc_min else thin
        right  = thick if col == calc_max else thin
        top    = thick if (col in calc_cols and row_idx == 1) else thin
        bottom = thick if (col in calc_cols and row_idx == ws.max_row) else thin
        return Border(left=left, right=right, top=top, bottom=bottom)

    # Kopfzeile
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = _FILL_HEADER
        cell.font   = _FONT_HEADER
        cell.border = cell_border(col, 1)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 22

    # Datenzeilen
    for row_idx in range(2, ws.max_row + 1):
        base_fill = _FILL_ZEBRA_E if row_idx % 2 == 0 else _FILL_ZEBRA_O

        # Service-Date-Pruefung: SD-Monat frueherer als Rechnungsmonat?
        sd_warn = False
        if sd_col and datum_col:
            sd_val    = ws.cell(row=row_idx, column=sd_col).value
            datum_val = ws.cell(row=row_idx, column=datum_col).value
            if sd_val and datum_val:
                try:
                    sd_dt    = datetime.strptime(str(sd_val).strip(),    "%d.%m.%Y")
                    datum_dt = datetime.strptime(str(datum_val).strip(), "%d.%m.%Y")
                    if (sd_dt.year, sd_dt.month) < (datum_dt.year, datum_dt.month):
                        sd_warn = True
                except ValueError:
                    pass

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font   = _FONT_BODY
            cell.border = cell_border(col, row_idx)
            cell.alignment = Alignment(vertical="center")

            # Zahlenformat
            if col in num_cols:
                cell.number_format = _NUM_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Farbe: Warnung > errechnete CZK-Spalten > Zebrastreifen
            if sd_warn and col == sd_col:
                cell.fill = _FILL_WARN
            elif col in calc_cols:
                cell.fill = _FILL_CALC
            else:
                cell.fill = base_fill

    # Spaltenbreiten automatisch anpassen (max. 40)
    for col in range(1, ws.max_column + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 40)

# ── CNB-Wechselkurse ─────────────────────────────────────────────────────
# FAW_NR-Mapping: Schluessel in 3216/3226 -> ISO-Waehrungscode
_FAW_MAP     = {1: "EUR", 22: "CZK", 23: "PLN", 24: "HUF"}
_CONVERT_FAW = {1, 23}   # 1=EUR, 23=PLN werden umgerechnet
_ALREADY_CZK = {22}      # 22=CZK: keine Umrechnung noetig, Betrag wird 1:1 uebernommen
                         # (24=HUF bleibt unangetastet, keine Umrechnung)

# Moegliche Spaltennamen fuer Fremdwaehrungsbetrag und Buchungstext
_FW_COL_NAMES   = {"Fremdwährung", "Fremdwaehrung"}
_TEXT_COL_NAMES = {"Text", "text", "Buchungstext", "Bezeichnung"}

_cnb_cache: dict = {}   # date_str -> (rates_dict, actual_date_str)

# Erkennt einen deutschen Dezimalwert am Textende, z.B. "Warenart XY 24,865"
_MANUAL_RATE_RE = re.compile(r'(\d{1,3}(?:\.\d{3})*,\d+)\s*$')


def _extract_manual_rate(text_val) -> float | None:
    """Gibt den manuell eingetragenen Kurs am Zeilenende zurueck, oder None."""
    if pd.isna(text_val):
        return None
    m = _MANUAL_RATE_RE.search(str(text_val).strip())
    if m:
        try:
            return float(m.group(1).replace(".", "").replace(",", "."))
        except ValueError:
            return None
    return None


def _fetch_cnb_rates(date_str: str) -> tuple[dict, str]:
    """
    Laedt CNB-Tageskurse fuer date_str (TT.MM.JJJJ).
    Faellt auf den naechsten verfuegbaren Handelstag zurueck (Wochenende/Feiertag).
    Gibt (rates_dict, tatsaechliches_datum) zurueck.
    """
    dt = datetime.strptime(date_str, "%d.%m.%Y")
    for offset in range(7):
        check_dt  = dt - timedelta(days=offset)
        check_str = check_dt.strftime("%d.%m.%Y")
        url = (
            "https://www.cnb.cz/en/financial-markets/foreign-exchange-market/"
            "central-bank-exchange-rate-fixing/central-bank-exchange-rate-fixing/"
            f"daily.txt?date={check_str}"
        )
        try:
            with urllib.request.urlopen(url, timeout=10) as resp:
                text = resp.read().decode("utf-8")
        except Exception as exc:
            raise RuntimeError(f"CNB nicht erreichbar ({check_str}): {exc}") from exc

        rates: dict = {}
        lines = text.strip().splitlines()
        for line in lines[2:]:          # erste zwei Zeilen: Datum-Zeile + Kopfzeile
            parts = line.split("|")
            if len(parts) == 5:
                try:
                    amount = int(parts[2])
                    code   = parts[3].strip()
                    rate   = float(parts[4].replace(",", "."))
                    rates[code] = rate / amount   # Kurs pro 1 Einheit
                except (ValueError, IndexError):
                    pass

        if rates:
            return rates, check_str

    raise RuntimeError(f"Keine CNB-Kurse fuer {date_str} nach 7 Rueckfalltagen gefunden.")


def _get_cnb_rate(date_str: str, currency: str) -> tuple[float | None, str]:
    """
    Gibt (Kurs, tatsaechliches_datum) fuer eine Waehrung zurueck.
    Ergebnis wird gecacht; None wenn Waehrung nicht gefunden.
    """
    if date_str not in _cnb_cache:
        rates, actual = _fetch_cnb_rates(date_str)
        _cnb_cache[date_str] = (rates, actual)
    rates, actual = _cnb_cache[date_str]
    return rates.get(currency), actual


def _apply_cnb_conversion(df: pd.DataFrame, source_file: str) -> list[str]:
    """
    3216: Kein automatischer CNB-Kurs. Steht ein manueller Kurs im Text, wird damit umgerechnet.
    3226: CNB-Kurs per Rechnungsdatum. Steht bereits ein manueller Kurs im Text, wird dieser
          vorrangig verwendet (tritt auf wenn Rechnung nicht in Kreditoren.xlsx gefunden wurde).
    Original Fremdwaehrung, deren Waehrungs-Spalte, Betrag und dessen Waehrungs-Spalte bleiben
    unveraendert. Zusaetzlich werden direkt daneben "Betrag CZK" und "Waehrung CZK" eingefuegt.
    Ist die Rechnung bereits in CZK (FAW_NR 22), wird der Betrag 1:1 uebernommen (keine Umrechnung
    noetig). HUF (FAW_NR 24) bleibt wie bisher unangetastet.
    Bei 3226 mit automatischem CNB-Kurs wird dieser zusaetzlich hinter die Warenart im Text
    angehaengt (nicht wenn bereits ein manueller Kurs im Text stand).
    """
    fw_col   = next((c for c in df.columns if c in _FW_COL_NAMES),   None)
    text_col = next((c for c in df.columns if c in _TEXT_COL_NAMES), None)
    faw_col  = "FAW_NR" if "FAW_NR" in df.columns else None

    is_3216 = source_file == "3216.xlsx"

    if not faw_col:
        return [f"[{source_file}] Spalte FAW_NR nicht gefunden – Umrechnung uebersprungen"]
    if not fw_col:
        return [f"[{source_file}] Keine Fremdwaehrungs-Spalte gefunden – Umrechnung uebersprungen"]

    # Neue Spalten hinter der Original-Waehrungsspalte einfuegen (nicht direkt hinter
    # Fremdwaehrung), damit die Original-Waehrungseinheit neben dem Original-Betrag stehen bleibt.
    _known_currencies = {"EUR", "PLN", "CZK", "HUF", "USD", "GBP", "CHF"}
    fw_pos = df.columns.get_loc(fw_col)
    insert_pos = fw_pos + 1
    if fw_pos + 1 < len(df.columns):
        next_col = df.columns[fw_pos + 1]
        vals = df[next_col].dropna().astype(str).str.strip().str.upper()
        if not vals.empty and vals.isin(_known_currencies).mean() > 0.5:
            insert_pos = fw_pos + 2

    df.insert(insert_pos, "Betrag CZK", None)
    df.insert(insert_pos + 1, "Währung CZK", None)

    conv_warnings: list[str] = []

    for idx in df.index:
        # FAW_NR -> Waehrungscode
        try:
            faw = int(float(str(df.at[idx, faw_col]).strip()))
        except (ValueError, TypeError):
            continue
        if faw not in _CONVERT_FAW and faw not in _ALREADY_CZK:
            continue  # z.B. HUF – bleibt unangetastet

        try:
            fw_val = float(df.at[idx, fw_col])
        except (ValueError, TypeError):
            conv_warnings.append(f"[{source_file}] Zeile {idx}: Kein gueltiger Betrag in {fw_col}")
            continue

        if faw in _ALREADY_CZK:
            # Bereits in CZK – keine Umrechnung noetig, 1:1 uebernehmen
            df.at[idx, "Betrag CZK"]  = round(fw_val, 2)
            df.at[idx, "Währung CZK"] = "CZK"
            continue

        currency  = _FAW_MAP[faw]
        datum_str = str(df.at[idx, "Datum"]).strip()

        # Manuellen Kurs aus Text lesen (falls vorhanden)
        manual_rate = _extract_manual_rate(df.at[idx, text_col]) if text_col else None

        append_rate = False  # Kurs hinter die Warenart im Text haengen (nur bei automatischem CNB-Kurs)

        if is_3216:
            # 3216: nur umrechnen wenn manueller Kurs im Text steht
            if manual_rate is None:
                continue
            rate = manual_rate
        else:
            # 3226: manueller Kurs hat Vorrang, sonst CNB per Rechnungsdatum
            if manual_rate is not None:
                rate = manual_rate
            else:
                try:
                    rate, actual_date = _get_cnb_rate(datum_str, currency)
                except RuntimeError as exc:
                    conv_warnings.append(f"[{source_file}] Zeile {idx}: {exc}")
                    continue
                if rate is None:
                    conv_warnings.append(
                        f"[{source_file}] Zeile {idx}: CNB kennt keinen Kurs fuer {currency} am {actual_date}"
                    )
                    continue
                if actual_date != datum_str:
                    log(f"  [{source_file}] {datum_str} kein Handelstag – CNB-Kurs von {actual_date}")
                append_rate = True

        # Umrechnung in neuer Spalte; Original-Fremdwaehrung/-Betrag bleibt unveraendert
        df.at[idx, "Betrag CZK"]  = round(fw_val * rate, 2)
        df.at[idx, "Währung CZK"] = "CZK"

        # Bei automatischem CNB-Kurs (3226): Kurs wie bisher hinter die Warenart im Text anhaengen
        if append_rate and text_col:
            rate_str = f"{rate:.3f}".replace(".", ",")
            existing = str(df.at[idx, text_col]) if pd.notna(df.at[idx, text_col]) else ""
            df.at[idx, text_col] = existing.rstrip() + f" {rate_str}"

    return conv_warnings


# Fix 1: Pfad relativ zum Skript, nicht absolut hardcoded
LOG_FILE = os.path.join(os.path.dirname(__file__), "build_log.txt")
_log = []

REQUIRED_COLS = {"Referenz", "Gegenkto", "Datum"}


def log(msg):
    _log.append(msg)
    print(msg.encode("ascii", "replace").decode())


def ask_folder():
    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title="Ordner mit 3216.xlsx und 3226.xlsx auswaehlen")
    root.destroy()
    return folder


# Fix 5: eigene Hilfsfunktion fuer Datei-Dialog (fuer Kreditoren-Fallback)
def ask_file(title, filetypes):
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()
    return path


def lookup_service_date(kred, ref, gegenkto):
    ref = str(ref).strip()
    gegenkto = str(gegenkto).strip().replace(".0", "")

    match = kred[(kred["Invoice Number"] == ref) & (kred["Company Account"] == gegenkto)]
    if len(match) == 1:
        return match.iloc[0]["Service Date"], "OK"

    ref_norm = ref.replace(" ", "")
    match = kred[(kred["Invoice Number normalized"] == ref_norm) & (kred["Company Account"] == gegenkto)]
    if len(match) == 1:
        found = match.iloc[0]["Invoice Number"]
        return match.iloc[0]["Service Date"], f"FUZZY ({ref} -> {found})"

    if len(match) == 0:
        return None, f"KEIN TREFFER: Ref={ref} Kred={gegenkto}"
    return None, f"MEHRDEUTIG: Ref={ref} Kred={gegenkto}"


def main():
    # Ordner auswaehlen
    folder = ask_folder()
    if not folder:
        print("Kein Ordner ausgewaehlt. Abbruch.")
        return

    log(f"Ordner: {folder}")

    # Pflichtdateien pruefen
    for fname in ["3216.xlsx", "3226.xlsx"]:
        if not os.path.exists(os.path.join(folder, fname)):
            messagebox.showerror("Fehler", f"{fname} nicht im Ordner gefunden.")
            return

    # Fix 5: Kreditoren laden — Fallback auf Datei-Dialog wenn Netzwerkpfad nicht erreichbar
    kred_path = KREDITOREN_PATH
    if not os.path.exists(kred_path):
        log(f"Kreditoren.xlsx nicht gefunden unter: {kred_path}")
        messagebox.showwarning(
            "Kreditoren.xlsx nicht gefunden",
            f"Die Datei wurde nicht gefunden:\n{kred_path}\n\nBitte im naechsten Schritt manuell auswaehlen."
        )
        kred_path = ask_file(
            title="Kreditoren.xlsx auswaehlen",
            filetypes=[("Excel-Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
        )
        if not kred_path:
            messagebox.showerror("Fehler", "Keine Kreditoren.xlsx ausgewaehlt. Abbruch.")
            return

    log("Lade Kreditoren.xlsx ...")
    kred = pd.read_excel(kred_path)
    kred["Invoice Number"] = kred["Invoice Number"].astype(str).str.strip()
    kred["Invoice Number normalized"] = kred["Invoice Number"].str.replace(" ", "", regex=False)
    kred["Company Account"] = kred["Company Account"].astype(str).str.strip().str.replace(".0", "", regex=False)

    sheets = {}
    warnings = []

    for source_file, sheet_name in [("3216.xlsx", "Reverse Charge - 3216"), ("3226.xlsx", "IGE - 3226")]:
        path = os.path.join(folder, source_file)
        df = pd.read_excel(path)
        df = df.dropna(subset=["Referenz", "Gegenkto"], how="all")

        # Fix 2: Spaltenvalidierung mit verstaendlicher Fehlermeldung
        missing = REQUIRED_COLS - set(df.columns)
        if missing:
            messagebox.showerror(
                "Spalten fehlen",
                f"{source_file}: Folgende Pflichtspalten fehlen:\n{', '.join(sorted(missing))}"
            )
            return

        service_dates = []
        for _, row in df.iterrows():
            sd, status = lookup_service_date(kred, row["Referenz"], row["Gegenkto"])
            if status != "OK":
                log(f"  [{source_file}] {status}")
                warnings.append(f"[{source_file}] {status}")
            if sd is not None and hasattr(sd, "strftime"):
                sd = sd.strftime("%d.%m.%Y")
            service_dates.append(sd)

        # Datum im deutschen Format
        df["Datum"] = pd.to_datetime(df["Datum"]).dt.strftime("%d.%m.%Y")

        # EUR/PLN -> CZK Umrechnung via CNB (vor Service-Date-Insert, damit Datum schon String ist)
        conv_warns = _apply_cnb_conversion(df, source_file)
        for w in conv_warns:
            log(f"  {w}")
            warnings.append(w)

        df.insert(1, "Service Date", service_dates)

        # Fix 4: Unnamed-Spalten anhand ihrer tatsaechlichen Position umbenennen
        cols = list(df.columns)
        for i, col in enumerate(cols):
            if col.startswith("Unnamed:"):
                cols[i] = f"Unnamed: {i}"
        df.columns = cols

        sheets[sheet_name] = df
        log(f"{source_file}: {len(df)} Zeilen verarbeitet")

    output_path = os.path.join(folder, "Reverse Charge and IGE.xlsx")

    # Fix 7: Ueberschreib-Schutz
    if os.path.exists(output_path):
        if not messagebox.askyesno(
            "Datei ueberschreiben?",
            f"Die Datei existiert bereits:\n{output_path}\n\nUeberschreiben?"
        ):
            log("Abbruch durch Benutzer (kein Ueberschreiben).")
            return

    # Fix 3: PermissionError abfangen (z.B. Datei noch in Excel geoeffnet)
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            # Formatierung nach dem Schreiben aller Blaetter anwenden
            for sheet_name, df in sheets.items():
                _apply_formatting(writer.sheets[sheet_name], df)
    except PermissionError:
        messagebox.showerror(
            "Datei gesperrt",
            f"Datei konnte nicht gespeichert werden.\n\nBitte zuerst schliessen:\n{output_path}"
        )
        return

    log(f"\nGespeichert: {output_path}")

    with open(LOG_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(_log))

    if warnings:
        msg = f"Datei gespeichert.\n\nHinweise ({len(warnings)}):\n" + "\n".join(warnings)
        messagebox.showwarning("Fertig mit Hinweisen", msg)
    else:
        messagebox.showinfo("Fertig", "Reverse Charge and IGE.xlsx wurde erfolgreich erstellt.")


if __name__ == "__main__":
    # Fix 6: Unerwartete Fehler abfangen und als Dialog anzeigen statt stilles Absturz
    try:
        main()
    except Exception as e:
        messagebox.showerror("Unerwarteter Fehler", str(e))
        sys.exit(1)
